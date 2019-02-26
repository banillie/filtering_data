'''

Programme for pulling out a single bit of data across chosen number of quarters. Data of interest is to be specified

It outputs a workbook with some conditional formatting to show, 1) changes in reported data, 2)when projects were
reporting.

'''
from openpyxl import Workbook
from bcompiler.utils import project_data_from_master
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import random



def data_return(project_list, data_key, dict_list):
    wb = Workbook()
    ws = wb.active

    '''lists project names in ws'''
    for x in range(0, len(project_list)):
        ws.cell(row=x + 2, column=1, value=project_list[x])

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=1).value
        print(project_name)
        col_start = 2
        for i, dictionary in enumerate(dict_list):
            if project_name in dictionary:
                ws.cell(row=row_num, column=col_start).value = dictionary[project_name][data_key]
                if dictionary[project_name][data_key] == None:
                    ws.cell(row=row_num, column=col_start).value = 'None'
                try:
                    if dict_list[i+1][project_name][data_key] != dictionary[project_name][data_key]:
                        ws.cell(row=row_num, column=col_start).font = red_text
                except (IndexError, KeyError):
                    pass
                col_start += 1
            else:
                ws.cell(row=row_num, column=col_start).value = 'Not reporting'
                col_start += 1

    quarter_labels = get_quarter_stamp(list_of_dicts)
    ws.cell(row=1, column=1, value='Project')
    for i, label in enumerate(quarter_labels):
        ws.cell(row=1, column=i + 2, value=label)

    conditional_formatting(ws)  # apply conditional formatting

    return wb

'''function for applying rag rating conditional formatting colouring if required'''
def conditional_formatting(worksheet):
    ag_text = Font(color="000000")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",A1)))']
    worksheet.conditional_formatting.add('A1:H50', rule)

    ar_text = Font(color="000000")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",A1)))']
    worksheet.conditional_formatting.add('A1:H50', rule)

    red_text = Font(color="000000")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",A1)))']
    worksheet.conditional_formatting.add('A1:H50', rule)

    green_text = Font(color="000000")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",A1)))']
    worksheet.conditional_formatting.add('A1:H50', rule)

    amber_text = Font(color="000000")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",A1)))']
    worksheet.conditional_formatting.add('A1:H50', rule)

    grey_text = Font(color="f0f0f0")
    grey_fill = PatternFill(bgColor="f0f0f0")
    dxf = DifferentialStyle(font=grey_text, fill=grey_fill)
    rule = Rule(type="containsText", operator="containsText", text="Not reporting", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Not reporting",A1)))']
    worksheet.conditional_formatting.add('A1:H50', rule)

    # highlighting new projects
    red_text = Font(color="000000")
    white_fill = PatternFill(bgColor="000000")
    dxf = DifferentialStyle(font=red_text, fill=white_fill)
    rule = Rule(type="containsText", operator="containsText", text="NEW", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("NEW",A1)))']
    worksheet.conditional_formatting.add('A1:H50', rule)

    return worksheet


def get_all_project_names(dict_list):
    output_list = []
    for dict in list_of_dicts:
        for name in dict:
            if name not in output_list:
                output_list.append(name)

    return output_list

def get_quarter_stamp(dict_list):
    output_list = []
    for dict in dict_list:
        proj_name = random.choice(list(dict.keys()))
        quarter_stamp = dict[proj_name]['Reporting period (GMPP - Snapshot Date)']
        output_list.append(quarter_stamp)

    return output_list

red_text = Font(color="FF0000")

'''master data sources'''
q3_1819 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2018.xlsx')
q2_1819 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx')
q1_1819 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_1_2018.xlsx')
q4_1718 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_4_2017.xlsx')
q3_1718 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2017.xlsx')
q2_1718 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2017.xlsx')
q1_1718 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_1_2017.xlsx')

list_of_dicts = [q3_1819, q2_1819, q1_1819, q4_1718, q3_1718, q2_1718, q1_1718]

'''compiling list of projects. there are two options'''
one_quarter_list = list(q3_1819.keys())
combined_quarters_list = get_all_project_names(list_of_dicts)

'''filter via group'''
# group_names = ['Rail Group', 'HSMRPG', 'International Security and Environment', 'Roads Devolution & Motoring']
# filtered_list_one = filter_group(one, 'HSMRPG')
# filtered_list_two = filter_group(two, 'HSMRPG')
# overall_list = sorted(list(set(filtered_list_one + filtered_list_two)))  ### To be completed

'''set data of interest'''
data_interest = 'Project stage'
# 'Overall Resource DCA - Now', 'Project Delivery - Now', 'Project MM18 Forecast - Actual',
# 'Project MM18 Original Baseline'  # project start date baseline , 'Overall Resource DCA - Now'
# 'Real or Nominal - Actual/Forecast', 'Total Forecast', 'Project Delivery - Now'

'''running programme'''
run = data_return(combined_quarters_list, data_interest, list_of_dicts)

run.save('C:\\Users\\Standalone\\Will\\project_stages.xlsx')

#tyr