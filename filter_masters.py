from openpyxl import load_workbook, Workbook
#from collections import OrderedDict
#from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font
import datetime
from bcompiler.utils import project_data_from_master


def first_function(Q2_master, other_master):
    ws = other_master.active

    project_names = Q2_master.keys()

    for col_num in range(2, ws.max_column+1):
        name_check = ws.cell(row=1, column=col_num).value
        if name_check in project_names:
            for row_num in range(2, ws.max_row + 1):
                key = ws.cell(row=row_num, column=1).value
                value = ws.cell(row=row_num, column=col_num).value

                try:
                    if type(Q2_master[name_check][key]) is datetime.date:
                        if type(value)is datetime.datetime:
                            if value.date() != Q2_master[name_check][key]:
                                ws.cell(row=row_num, column=col_num).font = blue_text
                    elif value == Q2_master[name_check][key]:
                        pass
                    else:
                        ws.cell(row=row_num, column=col_num).font = blue_text
                except KeyError:
                    ws.cell(row=row_num, column=col_num).font = blue_text

    return other_master

def print_text(blue_workbook, red_workbook):
    ws_blue = blue_workbook.active
    ws_red = red_workbook.active
    no = 0
    for col_num in range(2, ws_red.max_column+1):
        print(ws_blue.cell(row=1, column=col_num))
        print(ws_red.cell(row=1, column=col_num))
        for row_num in range(2, ws_red.max_row+1):
            if ws_blue.cell(row=row_num, column=col_num).font.color.rgb == '000000FF':  # blue
                if ws_red.cell(row=row_num, column=col_num).font.color.rgb == 'FFFC2525': # red
                    pass
                else:
                    new_value = ws_blue.cell(row=row_num, column=col_num).value
                    print(new_value)
                    ws_red.cell(row=row_num, column=col_num).value = new_value
                    ws_red.cell(row=row_num, column=col_num).font = blue_text

    return red_workbook



blue_text = Font(color="0000FF")

with_red = load_workbook("C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx")
with_blue = load_workbook("C:\\Users\\Standalone\\Will\\masters folder\\test_merge.xlsx")

Q2_one = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx")
Q2_two = load_workbook("C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018_to_clarify.xlsx")

b = print_text(with_blue, with_red)
b.save("C:\\Users\\Standalone\\Will\\masters folder\\merged.xlsx")

#a = first_function(Q2_one, Q2_two)

#a.save("C:\\Users\\Standalone\\Will\\masters folder\\test_merge.xlsx") hello